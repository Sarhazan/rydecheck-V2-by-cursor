import os
import requests
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_FILE = os.path.join(os.path.dirname(BASE_DIR), "דוח גט 1125.xlsx")
UPLOADED_FILE = os.path.join(BASE_DIR, "uploads", "supplier2_דוח גט 1125.xlsx")


def count_rows(path: str) -> int:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    count_d = 0
    for row_num in range(16, 642):  # user expectation 16-641
        cell = ws.cell(row=row_num, column=4)
        if cell.value is not None:
            count_d += 1
    max_row = ws.max_row
    wb.close()
    return max_row, count_d


def upload_gett() -> str:
    url = "http://localhost:5000/api/upload"
    with open(ROOT_FILE, "rb") as f:
        files = {"file": (os.path.basename(ROOT_FILE), f, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        data = {"type": "supplier2"}
        resp = requests.post(url, files=files, data=data)
    print("upload status:", resp.status_code)
    print("upload resp:", resp.text[:400])
    resp.raise_for_status()
    return resp.json()["filename"]


def call_gett_columns(path: str):
    resp = requests.post("http://localhost:5000/api/gett-columns", json={"filename": path})
    print("gett-columns status:", resp.status_code)
    print("gett-columns body:", resp.text[:400])
    if resp.ok:
        data = resp.json()
        print("gett-columns row_count:", data.get("row_count"))


def main():
    print("Root file:", ROOT_FILE)
    print("Exists root:", os.path.exists(ROOT_FILE))
    if os.path.exists(ROOT_FILE):
        max_row, count_d = count_rows(ROOT_FILE)
        print(f"Root: max_row={max_row}, rows16-641_colD={count_d}")

    print("\nUploading root file to API...")
    uploaded_path = upload_gett()
    print("Uploaded path:", uploaded_path)

    # Count uploaded file from disk
    if os.path.exists(uploaded_path):
        max_row_u, count_d_u = count_rows(uploaded_path)
        print(f"Uploaded file: max_row={max_row_u}, rows16-641_colD={count_d_u}")
    else:
        print("Uploaded file not found on disk.")

    # Call gett-columns for uploaded file
    call_gett_columns(uploaded_path)


if __name__ == "__main__":
    main()

