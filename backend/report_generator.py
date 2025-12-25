from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List
import os
import tempfile
from datetime import datetime

class ReportGenerator:
    """Generate Excel reports for department cost allocations"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def generate_department_report(self, department_name: str, 
                                  department_data: Dict[str, Any],
                                  output_path: str = None) -> str:
        """
        Generate Excel report for a single department
        
        Args:
            department_name: Name of the department
            department_data: Department allocation data from DepartmentAllocator
            output_path: Optional path to save file (if None, creates temp file)
        
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = department_name[:31]  # Excel sheet name limit
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        # Summary section
        row = 1
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = f'דוח עלויות מחלקה: {department_name}'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Summary statistics
        ws[f'A{row}'] = 'סה"כ עלות'
        ws[f'B{row}'] = f'₪{department_data["total_cost"]:,.2f}'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font(bold=True, size=11)
        row += 1
        
        ws[f'A{row}'] = 'מספר נסיעות'
        ws[f'B{row}'] = department_data['ride_count']
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        avg_cost = department_data['total_cost'] / department_data['ride_count'] if department_data['ride_count'] > 0 else 0
        ws[f'A{row}'] = 'ממוצע עלות לנסיעה'
        ws[f'B{row}'] = f'₪{avg_cost:,.2f}'
        ws[f'A{row}'].font = Font(bold=True)
        row += 3
        
        # Detail table header
        headers = ['מזהה נסיעה', 'תאריך', 'שעה', 'מוצא', 'יעד', 'נוסעים', 'עובדים', 'עלות מוקצית', 'אחוז הקצאה']
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            col += 1
        
        row += 1
        
        # Detail rows
        for ride_detail in department_data.get('rides', []):
            ride = ride_detail.get('ride', {})
            employees = ride_detail.get('employees', [])
            allocation = ride_detail.get('allocation', 0)
            percentage = ride_detail.get('percentage', 0)
            
            # Collect employee names
            emp_names = [f"{emp.get('first_name', '')} {emp.get('last_name', '')}".strip() 
                        for emp in employees]
            emp_names_str = ', '.join(emp_names) if emp_names else ''
            
            # Collect passenger names from ride
            passengers = ride.get('passengers', [])
            passengers_str = ', '.join(passengers) if passengers else ''
            
            ws.cell(row=row, column=1, value=ride.get('trip_id', ''))
            ws.cell(row=row, column=2, value=ride.get('date', ''))
            ws.cell(row=row, column=3, value=ride.get('time', ''))
            ws.cell(row=row, column=4, value=ride.get('source', ''))
            ws.cell(row=row, column=5, value=ride.get('destination', ''))
            ws.cell(row=row, column=6, value=passengers_str)
            ws.cell(row=row, column=7, value=emp_names_str)
            ws.cell(row=row, column=8, value=f'₪{allocation:,.2f}')
            ws.cell(row=row, column=9, value=f'{percentage:.1f}%')
            
            # Set alignment
            for col in range(1, 10):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal='right' if col <= 7 else 'center',
                    vertical='center',
                    wrap_text=True
                )
            
            row += 1
        
        # Total row
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = 'סה"כ'
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=8, value=f'₪{department_data["total_cost"]:,.2f}')
        ws.cell(row=row, column=8).font = Font(bold=True)
        ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column widths
        for col in range(1, 10):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Set output path
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"department_report_{department_name}_{timestamp}.xlsx"
            filename = filename.replace('/', '_').replace('\\', '_')  # Safe filename
            output_path = os.path.join(self.temp_dir, filename)
        
        wb.save(output_path)
        return output_path
    
    def generate_all_departments_report(self, department_allocations: Dict[str, Any],
                                       output_path: str = None) -> str:
        """
        Generate summary report for all departments
        
        Args:
            department_allocations: Complete allocation data from DepartmentAllocator
            output_path: Optional path to save file
        
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "סיכום מחלקות"
        
        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        row = 1
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = 'סיכום עלויות לפי מחלקה'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Summary table header
        headers = ['מחלקה', 'סה"כ עלות', 'מספר נסיעות', 'ממוצע עלות לנסיעה']
        col = 1
        for header in headers:
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            col += 1
        
        row += 1
        
        # Summary rows
        dept_allocs = department_allocations.get('department_allocations', {})
        total_cost = 0
        total_rides = 0
        
        for dept_name, dept_data in sorted(dept_allocs.items()):
            total_cost += dept_data['total_cost']
            total_rides += dept_data['ride_count']
            avg_cost = dept_data['total_cost'] / dept_data['ride_count'] if dept_data['ride_count'] > 0 else 0
            
            ws.cell(row=row, column=1, value=dept_name)
            ws.cell(row=row, column=2, value=f'₪{dept_data["total_cost"]:,.2f}')
            ws.cell(row=row, column=3, value=dept_data['ride_count'])
            ws.cell(row=row, column=4, value=f'₪{avg_cost:,.2f}')
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = Alignment(
                    horizontal='right' if col == 1 else 'center',
                    vertical='center'
                )
            
            row += 1
        
        # Total row
        row += 1
        ws.cell(row=row, column=1, value='סה"כ כללי')
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=f'₪{total_cost:,.2f}')
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=3, value=total_rides)
        ws.cell(row=row, column=3).font = Font(bold=True)
        avg_total = total_cost / total_rides if total_rides > 0 else 0
        ws.cell(row=row, column=4, value=f'₪{avg_total:,.2f}')
        ws.cell(row=row, column=4).font = Font(bold=True)
        
        # Handle unassigned
        unassigned = department_allocations.get('unassigned', {})
        if unassigned.get('ride_count', 0) > 0:
            row += 2
            ws.cell(row=row, column=1, value='לא משויך למחלקה')
            ws.cell(row=row, column=1).font = Font(bold=True, italic=True)
            ws.cell(row=row, column=2, value=f'₪{unassigned["total_cost"]:,.2f}')
            ws.cell(row=row, column=3, value=unassigned['ride_count'])
        
        # Auto-adjust column widths
        for col in range(1, 5):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Set output path
        if not output_path:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"all_departments_summary_{timestamp}.xlsx"
            output_path = os.path.join(self.temp_dir, filename)
        
        wb.save(output_path)
        return output_path

